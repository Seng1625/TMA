from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Any, Dict, List

from flask import Flask, jsonify, render_template, request, send_from_directory

# NOTE: this app is meant for LOCAL use (prototype).
# It saves results to an Excel file (.xlsx) in the ./data folder.

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)


@app.get("/")
def home():
    return render_template("prototype.html")


def _safe_filename(name: str) -> str:
    name = re.sub(r"[^a-zA-Z0-9._-]+", "_", name).strip("._-")
    return name or "results"


@app.post("/api/save")
def save():
    payload: Dict[str, Any] = request.get_json(force=True, silent=False)  # type: ignore
    participant_id = str(payload.get("participantId", "P00000"))
    session_started = str(payload.get("sessionStartedAt", datetime.utcnow().isoformat()))
    results: List[Dict[str, Any]] = payload.get("results", []) or []

    # Lazy import
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    headers = [
        "participantId",
        "sessionStartedAt",
        "trialIndex",
        "trialId",
        "type",
        "title",
        "prompt",
        "options",
        "correct",
        "aiSuggestion",
        "aiConfidence",
        "aiIsCorrect",
        "initialChoice",
        "initialRTms",
        "initialCorrect",
        "followedAI_initial",
        "finalChoice",
        "finalRTms",
        "finalCorrect",
        "followedAI_final",
        "changed",
        "stimulusParams",
        "timestamp",
    ]
    ws.append(headers)

    import json
    for r in results:
        row = []
        for h in headers:
            v = r.get(h, "")
            if h in ("options", "stimulusParams"):
                v = json.dumps(v, ensure_ascii=False)
            row.append(v)
        ws.append(row)

    ws2 = wb.create_sheet("summary")
    ws2.append(["participantId", participant_id])
    ws2.append(["sessionStartedAt", session_started])
    ws2.append(["trials", len(results)])

    if results:
        changed = sum(1 for r in results if r.get("changed"))
        follow_init = sum(1 for r in results if r.get("followedAI_initial"))
        follow_final = sum(1 for r in results if r.get("followedAI_final"))
        init_acc = sum(1 for r in results if r.get("initialCorrect"))
        final_acc = sum(1 for r in results if r.get("finalCorrect"))

        ws2.append(["%changed", round(100 * changed / len(results), 1)])
        ws2.append(["%followAI_initial", round(100 * follow_init / len(results), 1)])
        ws2.append(["%followAI_final", round(100 * follow_final / len(results), 1)])
        ws2.append(["%accuracy_initial", round(100 * init_acc / len(results), 1)])
        ws2.append(["%accuracy_final", round(100 * final_acc / len(results), 1)])

    for sheet in (ws, ws2):
        for col in range(1, sheet.max_column + 1):
            max_len = 10
            for row in range(1, min(sheet.max_row, 200) + 1):
                val = sheet.cell(row=row, column=col).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val))[:80])
            sheet.column_dimensions[get_column_letter(col)].width = min(55, max_len + 2)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = _safe_filename(f"results_{participant_id}_{ts}.xlsx")
    filepath = os.path.join(DATA_DIR, filename)
    wb.save(filepath)

    return jsonify({
        "ok": True,
        "filename": filename,
        "download_url": f"/download/{filename}",
        "saved_to": filepath,
    })


@app.get("/download/<path:filename>")
def download(filename: str):
    return send_from_directory(DATA_DIR, filename, as_attachment=True)


if __name__ == "__main__":
    # Cloud-friendly: listen on all interfaces and use $PORT if provided (Render/Railway).
    port = int(os.environ.get("PORT", "8000"))
    app.run(host="0.0.0.0", port=port, debug=False)
